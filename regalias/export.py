"""
export.py — Generación del reporte Excel global y consolidado.

Cubre las tres fuentes que aparecen en la app:
  · Departamento  (tabla MatrizSeguimientoEvaluacion)
  · Descentralizadas (tabla OtrosEjecutoresDescentralizadas)
  · Municipios   (tabla OtrosEjecutoresMunicipios)

Cada hoja replica con fidelidad la información que el usuario ve en pantalla:
unidades correctas (días/meses), porcentajes para avance, mensajes de
semáforo, e incluso el "Reporte semanal de alertas" calculado en línea.

El reporte es independiente del filtro activo en el panel: siempre incluye
las tres fuentes si están disponibles.
"""
from regalias.constants import C, INTERVALOS, SEMAFOROS, COLS_EVAL, COLS_EVAL_LABELS
import polars as pl
import io
import datetime as _dt
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Paleta y mapeos de semáforo
# ─────────────────────────────────────────────────────────────────────────────
AZUL_OSC  = "003D6C"
AZUL_MED  = "1754AB"
VERDE_OSC = "005931"
GRIS_BRD  = "D1D5DB"
GRIS_ALT  = "F7FAFD"
BLANCO    = "FFFFFF"

# Relleno (background, foreground) por código de clasificación
SEM_FILL = {
    # Hitos 1, 2, 5
    "0-100":   ("D1FAE5", "065F46"),
    "101-150": ("FEF3C7", "92400E"),
    "151-180": ("FFEDD5", "9A3412"),
    ">180":    ("1E293B", "F1F5F9"),
    # Hito 3
    "0-15":    ("D1FAE5", "065F46"),
    "16-30":   ("FEF3C7", "92400E"),
    "31-45":   ("FFEDD5", "9A3412"),
    ">45":     ("1E293B", "F1F5F9"),
    # Hito 5 (meses)
    "0-1":     ("D1FAE5", "065F46"),
    "1.1-3":   ("FEF3C7", "92400E"),
    "3.1-6":   ("FFEDD5", "9A3412"),
    ">6":      ("1E293B", "F1F5F9"),
}

SEM_NIVEL = {
    "0-100": "Verde", "101-150": "Naranja", "151-180": "Rojo", ">180": "Negro",
    "0-15":  "Verde", "16-30":   "Naranja", "31-45":   "Rojo", ">45":    "Negro",
    "0-1":   "Verde", "1.1-3":   "Naranja", "3.1-6":   "Rojo", ">6":     "Negro",
}

# Mensajes (toma el primero que vea — los hitos comparten claves "0-100" etc.,
# pero el mensaje cambia. Esto es un compromiso aceptable porque el usuario
# verá el comentario contextual en la hoja "Detalle" donde sí mapeamos por hito).
SEM_MSG = {}
for hk, vals in SEMAFOROS.items():
    for label, (_, _, msg) in vals.items():
        SEM_MSG.setdefault(label, msg)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de formato
# ─────────────────────────────────────────────────────────────────────────────
def _border():
    s = Side(style="thin", color=GRIS_BRD)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, color="1A2332", size=11, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _align(h="left", wrap=True, v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_cell(cell, text):
    cell.value = text
    cell.font = Font(name="Calibri", bold=True, color=BLANCO, size=11)
    cell.fill = _fill(AZUL_OSC)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

def _data_cell(cell, value, bg=BLANCO, bold=False, color="1A2332",
               center=False, fmt=None, italic=False):
    cell.value = value
    cell.font = _font(bold=bold, color=color, italic=italic)
    cell.fill = _fill(bg)
    cell.alignment = _align("center" if center else "left", wrap=True)
    cell.border = _border()
    if fmt:
        cell.number_format = fmt

def _sem_cell(cell, clasi, bg_default=BLANCO, hito_col=None):
    """
    Pinta una celda con color de semáforo según la clasificación.
    Adjunta un comentario con el mensaje del hito correspondiente. Si se pasa
    `hito_col`, usa el mensaje específico del hito (más preciso para Detalle).
    """
    s = str(clasi) if clasi and str(clasi) not in ("None", "nan", "") else None
    if s and s in SEM_FILL:
        bg, fg = SEM_FILL[s]
        cell.value = s
        cell.font = Font(name="Calibri", bold=True, color=fg, size=11)
        cell.fill = _fill(bg)
        cell.alignment = _align("center")
        cell.border = _border()
        # Mensaje específico por hito si lo pedimos
        msg = ""
        if hito_col and hito_col in SEMAFOROS and s in SEMAFOROS[hito_col]:
            msg = SEMAFOROS[hito_col][s][2]
        else:
            msg = SEM_MSG.get(s, "")
        nivel = SEM_NIVEL.get(s, "")
        if msg:
            cell.comment = Comment(f"● {nivel}\n{msg}", "Sistema", height=70, width=320)
    else:
        _data_cell(cell, "—", bg=bg_default, center=True, color="9CA3AF")

def _title_row(ws, title, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title)
    c.font = Font(name="Calibri", bold=True, size=14, color=BLANCO)
    c.fill = _fill(AZUL_OSC)
    c.alignment = _align("left")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(2, 1, sub)
    c2.font = Font(name="Calibri", size=10, color=BLANCO, italic=True)
    c2.fill = _fill(VERDE_OSC)
    c2.alignment = _align("left")
    ws.row_dimensions[2].height = 18

def _avance_pct(v):
    """Normaliza avance: si viene como 0-1 lo escala a porcentaje 0-100."""
    if v is None or str(v) in ("None", "nan", ""):
        return None
    try:
        fv = float(v)
        return fv * 100.0 if fv <= 1.0001 else fv
    except Exception:
        return None

def _clasificar_h(hito_col, dias_num):
    """Clasifica un valor (en días) según los intervalos del hito.
    H5 se evalúa en meses (días/30) — el resto en días."""
    if dias_num is None:
        return None
    if hito_col == "hito_5_val":
        m = dias_num / 30.0
        if m <= 1: return "0-1"
        if m <= 3: return "1.1-3"
        if m <= 6: return "3.1-6"
        return ">6"
    for label, lo, hi in INTERVALOS.get(hito_col, []):
        if (hi is None and dias_num >= lo) or (hi is not None and lo <= dias_num <= hi):
            return label
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Reporte semanal — configuración compartida (Departamento + Descentralizadas)
# ─────────────────────────────────────────────────────────────────────────────
ALERTAS_NRN = {
    "hito_1_val": ["101-150", "151-180", ">180"],
    "hito_2_val": ["101-150", "151-180", ">180"],
    "hito_3_val": ["16-30", "31-45", ">45"],
    "hito_5_val": ["1.1-3", "3.1-6", ">6"],
}

REPORTE_CONFIG = [
    {"estado": "SIN CONTRATAR", "label": "Sin contratar",
     "hitos": [("clasi_1", "hito_1_val"), ("clasi_2", "hito_2_val")]},
    {"estado": "CONTRATADO SIN ACTA DE INICIO", "label": "Contratado sin acta de inicio",
     "hitos": [("clasi_3", "hito_3_val")]},
    {"estado": "CONTRATADO EN EJECUCIÓN", "label": "Contratado en ejecución",
     "hitos": [("clasi_5", "hito_5_val")]},
]

def _comentario_alerta(estado_up, conteos, n_total):
    """Comentario corto sobre el estado y el conteo de alertas."""
    n_alerta = sum(conteos.values())
    if n_alerta == 0:
        return "Ningún proyecto presenta alertas en este estado."
    pct = round(n_alerta / n_total * 100) if n_total else 0
    partes = [f"{n_alerta} de {n_total} proyecto(s) ({pct}%) con alertas que requieren atención."]
    if estado_up == "SIN CONTRATAR":
        n_negro = conteos.get(">180", 0)
        if n_negro:
            partes.append(f"{n_negro} en alerta negra (más de 180 días sin avance).")
    elif estado_up == "CONTRATADO SIN ACTA DE INICIO":
        n_negro = conteos.get(">45", 0)
        if n_negro:
            partes.append(f"{n_negro} superan los 45 días sin acta de inicio.")
    elif estado_up == "CONTRATADO EN EJECUCIÓN":
        n_negro = conteos.get(">6", 0)
        if n_negro:
            partes.append(f"{n_negro} con horizonte vencido más de 6 meses.")
    return " ".join(partes)


# ─────────────────────────────────────────────────────────────────────────────
# Hojas — helpers comunes
# ─────────────────────────────────────────────────────────────────────────────
DATE_COLS_DPTO = {
    "FECHA APROBACIÓN PROYECTO", "FECHA DE APERTURA DEL PRIMER PROCESO",
    "FECHA DE SUSCRIPCIÓN DEL CONTRATO PRINCIPAL", "FECHA ACTA INICIO", "HORIZONTE DEL PROYECTO",
    "FECHA DE FINALIZACIÓN", "FECHA DE CORTE GESPROY",
}
DATE_COLS_DESCENT = {
    "FECHA APROBACIÓN PROYECTO", "FECHA DE APERTURA DEL PRIMER PROCESO",
    "FECHA DE SUSCRIPCIÓN DEL CONTRATO PRINCIPAL", "FECHA ACTA INICIO", "HORIZONTE DEL PROYECTO",
    "FECHA DE CORTE GESPROY",
}
HITO_COLS_DPTO    = {"hito_1_val", "hito_2_val", "hito_3_val", "hito_4_val", "hito_5_val", "hito_6_val"}
HITO_COLS_DESCENT = {"hito_1_val", "hito_2_val", "hito_3_val", "hito_5_val"}
CLASI_COLS = {"clasi_1", "clasi_2", "clasi_3", "clasi_5", "clasi_6"}
HITO_BY_CLASI = {"clasi_1":"hito_1_val", "clasi_2":"hito_2_val",
                 "clasi_3":"hito_3_val", "clasi_5":"hito_5_val",
                 "clasi_6":"hito_6_val"}
PCT_COLS = {"AVANCE FISICO", "AVANCE FÍSICO", "AVANCE FINANCIERO"}

def _write_date(cell, v, bg):
    if v is not None and str(v) not in ("None", "nan", "NaT", ""):
        if isinstance(v, (_dt.date, _dt.datetime)):
            cell.value = v
            cell.number_format = "DD/MM/YYYY"
            cell.font = _font()
            cell.fill = _fill(bg)
            cell.alignment = _align("center")
            cell.border = _border()
        else:
            _data_cell(cell, str(v), bg=bg, center=True)
    else:
        _data_cell(cell, "—", bg=bg, center=True, color="9CA3AF")


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 1 · Resumen Departamento (entidad × hitos × suspendidos × cierre × total)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_resumen_dpto(wb, df_agr):
    ws = wb.active
    ws.title = "Resumen Departamento"
    ws.sheet_view.showGridLines = False

    cols = [
        ("Entidad / Secretaría",                30),
        ("H1 días\nSin contratar\nsin apertura", 13),
        ("H1 alerta",                            10),
        ("H2 días\nSin contratar\ncon apertura", 13),
        ("H2 alerta",                            10),
        ("H3 días\nContratado\nsin acta",        13),
        ("H3 alerta",                            10),
        ("H4 días\nEn ejecución",                13),
        ("H5 meses\nEn ejecución\nrezagado",     13),
        ("H5 alerta",                            10),
        ("H6 días\nTerminados",                  13),
        ("H6 alerta",                            10),
        ("H7\nSuspendidos\nen ejecución",        13),
        ("H8 días\nPara cierre",                 13),
        ("Suspendidos",                          11),
        ("Para cierre",                          11),
        ("Total",                                 9),
    ]
    NCOLS = len(cols)
    _title_row(ws,
               "Resumen por Entidad — Departamento de Sucre",
               f"Generado: {date.today().strftime('%d/%m/%Y')} · "
               "Promedios por hito (H5 en meses) · Niveles de alerta como semáforo",
               NCOLS)
    ws.row_dimensions[3].height = 6
    for i, (label, w) in enumerate(cols, 1):
        _header_cell(ws.cell(4, i), label)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 56

    if df_agr is None or df_agr.height == 0:
        ws.cell(5, 1, "Sin datos disponibles").font = _font(italic=True, color="9CA3AF")
        return

    # Hitos con semáforo (clasi): H1, H2, H3, H5, H6
    HITO_KEYS = [
        ("Hito 1 (días)", "hito_1_val"),
        ("Hito 2 (días)", "hito_2_val"),
        ("Hito 3 (días)", "hito_3_val"),
        ("Hito 5 (días)", "hito_5_val"),
        ("Hito 6 (días)", "hito_6_val"),
    ]

    for r_i, row in enumerate(df_agr.to_dicts(), 5):
        bg = GRIS_ALT if r_i % 2 == 0 else BLANCO
        ws.row_dimensions[r_i].height = 22
        ent = row.get("ENTIDAD O SECRETARIA") or ""
        _data_cell(ws.cell(r_i, 1), ent, bg=bg, bold=True, color=AZUL_MED)
        col = 2
        # H1, H2, H3 — valor + semáforo
        for src_col, hito_col in HITO_KEYS[:3]:
            val = row.get(src_col)
            num = round(float(val), 1) if val is not None and str(val) != "nan" else None
            _data_cell(ws.cell(r_i, col), num, bg=bg, center=True, fmt="#,##0.0")
            col += 1
            clasi = _clasificar_h(hito_col, num) if num is not None else None
            _sem_cell(ws.cell(r_i, col), clasi, bg_default=bg, hito_col=hito_col)
            col += 1
        # H4 — En ejecución (sin semáforo): solo valor
        val4 = row.get("Hito 4 (días)")
        num4 = round(float(val4), 1) if val4 is not None and str(val4) != "nan" else None
        _data_cell(ws.cell(r_i, col), num4, bg=bg, center=True, fmt="#,##0.0")
        col += 1
        # H5 (meses) + H6 — valor + semáforo
        for src_col, hito_col in HITO_KEYS[3:]:
            val = row.get(src_col)
            num = round(float(val), 1) if val is not None and str(val) != "nan" else None
            shown = round(num / 30.0, 1) if (hito_col == "hito_5_val" and num is not None) else num
            _data_cell(ws.cell(r_i, col), shown, bg=bg, center=True, fmt="#,##0.0")
            col += 1
            clasi = _clasificar_h(hito_col, num) if num is not None else None
            _sem_cell(ws.cell(r_i, col), clasi, bg_default=bg, hito_col=hito_col)
            col += 1
        # H7 — Proyectos suspendidos en ejecución (conteo, sin semáforo)
        v7  = row.get("Hito 7")
        iv7 = int(v7) if v7 is not None and str(v7) != "nan" else 0
        _data_cell(ws.cell(r_i, col), iv7, bg=bg, center=True)
        col += 1
        # H8 — Proyecto para cierre (promedio de días, sin semáforo)
        v8   = row.get("Hito 8 (días)")
        num8 = round(float(v8), 1) if v8 is not None and str(v8) != "nan" else None
        _data_cell(ws.cell(r_i, col), num8, bg=bg, center=True, fmt="#,##0.0")
        col += 1
        for extra in ("Suspendidos", "Para cierre", "Total"):
            v = row.get(extra)
            iv = int(v) if v is not None and str(v) != "nan" else 0
            if extra == "Total":
                _data_cell(ws.cell(r_i, col), iv, bg="EFF6FF",
                           center=True, bold=True, color=AZUL_MED)
            else:
                _data_cell(ws.cell(r_i, col), iv, bg=bg, center=True)
            col += 1


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 2 · Detalle Departamento (proyecto × fechas × hitos × alertas × mensajes)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_detalle_dpto(wb, df_f):
    ws = wb.create_sheet("Detalle Departamento")
    ws.sheet_view.showGridLines = False

    base = [
        ("Entidad / Secretaría",  26, "ENTIDAD O SECRETARIA"),
        ("BPIN",                  14, "BPIN"),
        ("Nombre del proyecto",   42, "NOMBRE PROYECTO"),
        ("Estado proyecto",       20, "ESTADO PROYECTO"),
        ("Estado contrato",       16, "ESTADO CONTRATO"),
    ]
    if "RESPONSABLE CARGUE EN GESPROY" in df_f.columns:
        base.append(("Responsable cargue\nGESPROY", 22, "RESPONSABLE CARGUE EN GESPROY"))
    if "AVANCE FISICO" in df_f.columns:
        base.append(("Avance\nfísico",     11, "AVANCE FISICO"))
    if "AVANCE FINANCIERO" in df_f.columns:
        base.append(("Avance\nfinanciero", 12, "AVANCE FINANCIERO"))
    base += [
        ("Fecha\naprobación",            12, "FECHA APROBACIÓN PROYECTO"),
        ("Fecha apertura\nprimer proceso", 14, "FECHA DE APERTURA DEL PRIMER PROCESO"),
        ("Fecha\nsuscripción",           12, "FECHA DE SUSCRIPCIÓN DEL CONTRATO PRINCIPAL"),
        ("Fecha acta\ninicio",           12, "FECHA ACTA INICIO"),
        ("Horizonte\nproyecto",          12, "HORIZONTE DEL PROYECTO"),
        ("Fecha\nfinalización",          12, "FECHA DE FINALIZACIÓN"),
        ("Fecha corte\nGESPROY",         12, "FECHA DE CORTE GESPROY"),
    ]
    # H1, H2, H3 — días + alerta + mensaje
    for n, etiqueta in [(1, "H1"), (2, "H2"), (3, "H3")]:
        base.append((f"{etiqueta}\ndías",       7, f"hito_{n}_val"))
        base.append((f"{etiqueta}\nalerta",     9, f"clasi_{n}"))
        base.append((f"{etiqueta} · Mensaje", 36, f"_msg_{n}"))
    # H4 — En ejecución (sin semáforo): solo días
    base.append(("H4\ndías\nEn ejecución", 9, "hito_4_val"))
    # H5 (meses) y H6 — días + alerta + mensaje
    for n, etiqueta in [(5, "H5"), (6, "H6")]:
        base.append((f"{etiqueta}\ndías",       7, f"hito_{n}_val"))
        base.append((f"{etiqueta}\nalerta",     9, f"clasi_{n}"))
        base.append((f"{etiqueta} · Mensaje", 36, f"_msg_{n}"))
    base.append(("Suspendido",   11, "_susp"))
    base.append(("Para cierre",  11, "_cierre"))

    cols = base
    NCOLS = len(cols)
    _title_row(ws,
               "Detalle de proyectos — Departamento de Sucre",
               f"Generado: {date.today().strftime('%d/%m/%Y')} · "
               "Fechas, días por hito y nivel de alerta · Hover sobre la celda de alerta para ver el mensaje",
               NCOLS)
    ws.row_dimensions[3].height = 6
    for i, (label, w, _) in enumerate(cols, 1):
        _header_cell(ws.cell(4, i), label)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 50

    rows = df_f.to_dicts()
    for r_i, row in enumerate(rows, 5):
        bg = GRIS_ALT if r_i % 2 == 0 else BLANCO
        ws.row_dimensions[r_i].height = 38
        # Pre-calcular mensajes por hito y flags
        msgs = {}
        # Solo hitos con semáforo (no incluye H4 = En ejecución).
        for n in (1, 2, 3, 5, 6):
            cv = row.get(f"clasi_{n}")
            cv_s = str(cv) if cv and str(cv) not in ("None", "nan", "") else None
            hk = HITO_BY_CLASI[f"clasi_{n}"]
            if cv_s and hk in SEMAFOROS and cv_s in SEMAFOROS[hk]:
                msgs[f"_msg_{n}"] = SEMAFOROS[hk][cv_s][2]
            else:
                msgs[f"_msg_{n}"] = ""
        susp_v   = row.get("Suspendidos")
        cierre_v = row.get("Para cierre")
        msgs["_susp"]   = "Sí" if (susp_v   not in (None, 0, "0") and str(susp_v)   != "nan") else "No"
        msgs["_cierre"] = "Sí" if (cierre_v not in (None, 0, "0") and str(cierre_v) != "nan") else "No"

        for c_i, (_, _, attr) in enumerate(cols, 1):
            cell = ws.cell(r_i, c_i)
            if attr in CLASI_COLS:
                _sem_cell(cell, row.get(attr), bg_default=bg, hito_col=HITO_BY_CLASI.get(attr))
            elif attr in HITO_COLS_DPTO:
                v = row.get(attr)
                num = round(float(v), 1) if v is not None and str(v) != "nan" else None
                _data_cell(cell, num, bg=bg, center=True, color="1A2332", fmt="#,##0.0")
            elif attr in PCT_COLS:
                pct = _avance_pct(row.get(attr))
                _data_cell(cell, round(pct, 1) if pct is not None else None,
                           bg=bg, center=True, fmt='0.0"%"')
            elif attr in DATE_COLS_DPTO:
                _write_date(cell, row.get(attr), bg)
            elif attr.startswith("_msg"):
                _data_cell(cell, msgs.get(attr, ""), bg=bg, color="374151", italic=True)
            elif attr in ("_susp", "_cierre"):
                v = msgs[attr]
                bg_flag = "FEF3C7" if v == "Sí" else bg
                color   = "92400E" if v == "Sí" else "6B7280"
                _data_cell(cell, v, bg=bg_flag, center=True, bold=(v == "Sí"), color=color)
            else:
                v = row.get(attr)
                _data_cell(cell, v if v else "—", bg=bg)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 3 / Hoja 6 · Reporte semanal de alertas
#                  (Departamento usa ENTIDAD O SECRETARIA, Descent. usa EJECUTOR)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_reporte_semanal(wb, df, col_agrup, label_agrup, sheet_name, sub_titulo):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    cols = [
        (label_agrup,                  30),
        ("Estado del proyecto",        28),
        ("N.° proyectos\ncon alerta",  14),
        ("De un total de",             14),
        ("Alertas (clasificación)",    32),
        ("Comentario",                 60),
    ]
    NCOLS = len(cols)
    _title_row(ws,
               f"Reporte semanal de alertas — {sub_titulo}",
               f"Generado: {date.today().strftime('%d/%m/%Y')} · "
               "Solo proyectos con semáforo naranja, rojo o negro",
               NCOLS)
    ws.row_dimensions[3].height = 6
    for i, (label, w) in enumerate(cols, 1):
        _header_cell(ws.cell(4, i), label)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 38

    if df is None or df.height == 0 or col_agrup not in df.columns:
        ws.cell(5, 1, "Sin datos disponibles").font = _font(italic=True, color="9CA3AF")
        return

    # Solo entidades/ejecutores presentes
    entidades = sorted([x for x in df[col_agrup].drop_nulls().unique().to_list() if x])
    r = 5
    sin_alertas = True
    for cfg in REPORTE_CONFIG:
        estado_up = cfg["estado"]
        df_estado = df.filter(pl.col("ESTADO PROYECTO") == estado_up)
        if df_estado.height == 0:
            continue
        # Encabezado del estado (banda destacada)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
        c = ws.cell(r, 1, f"  {cfg['label'].upper()}  ·  {df_estado.height} proyecto(s) en este estado")
        c.font = Font(name="Calibri", bold=True, color=AZUL_OSC, size=10)
        c.fill = _fill("E8F0F9")
        c.alignment = _align("left")
        c.border = _border()
        ws.row_dimensions[r].height = 22
        r += 1

        for ent in entidades:
            df_ent = df_estado.filter(pl.col(col_agrup) == ent)
            if df_ent.height == 0:
                continue
            conteos = {}
            for clasi_col, hito_col in cfg["hitos"]:
                if clasi_col not in df.columns:
                    continue
                for alerta in ALERTAS_NRN.get(hito_col, []):
                    n = int(df_ent.filter(pl.col(clasi_col) == alerta).height)
                    if n > 0:
                        conteos[alerta] = conteos.get(alerta, 0) + n
            n_alerta = sum(conteos.values())
            if n_alerta == 0:
                continue
            sin_alertas = False
            bg = GRIS_ALT if r % 2 == 0 else BLANCO
            ws.row_dimensions[r].height = 36
            _data_cell(ws.cell(r, 1), ent, bg=bg, bold=True, color=AZUL_MED)
            _data_cell(ws.cell(r, 2), cfg["label"], bg=bg, color=AZUL_OSC)
            _data_cell(ws.cell(r, 3), n_alerta, bg=bg, center=True, bold=True, color=AZUL_OSC)
            _data_cell(ws.cell(r, 4), df_ent.height, bg=bg, center=True, color="6B7280")
            alertas_str = ", ".join(
                f"{k} ({v})" for k, v in sorted(conteos.items(), key=lambda x: -x[1])
            )
            _data_cell(ws.cell(r, 5), alertas_str, bg=bg, color="374151")
            _data_cell(ws.cell(r, 6),
                       _comentario_alerta(estado_up, conteos, df_ent.height),
                       bg=bg, color="374151")
            r += 1

    if sin_alertas:
        ws.cell(r, 1, "No se encontraron proyectos con alertas naranja, roja o negra.").font = _font(italic=True, color="9CA3AF")


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 4 · Resumen Descentralizadas (hitos 1-4, sin H5)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_resumen_descent(wb, df_descent):
    ws = wb.create_sheet("Resumen Descentralizadas")
    ws.sheet_view.showGridLines = False

    cols = [
        ("Ejecutor",                             32),
        ("H1 días\nSin contratar\nsin apertura", 13),
        ("H1 alerta",                            10),
        ("H2 días\nSin contratar\ncon apertura", 13),
        ("H2 alerta",                            10),
        ("H3 días\nContratado\nsin acta",        13),
        ("H3 alerta",                            10),
        ("H5 meses\nEn ejecución\nrezagado",     13),
        ("H5 alerta",                            10),
        ("Suspendidos",                          11),
        ("Para cierre",                          11),
        ("Total",                                 9),
    ]
    NCOLS = len(cols)
    _title_row(ws,
               "Resumen por Ejecutor — Descentralizadas",
               f"Generado: {date.today().strftime('%d/%m/%Y')} · "
               "Hitos 1, 2, 3 y 5 (H4 requiere contratos · H6 no aplica: sin fecha de finalización)",
               NCOLS)
    ws.row_dimensions[3].height = 6
    for i, (label, w) in enumerate(cols, 1):
        _header_cell(ws.cell(4, i), label)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 56

    if df_descent is None or df_descent.height == 0:
        ws.cell(5, 1, "Sin datos disponibles").font = _font(italic=True, color="9CA3AF")
        return

    hito_cols = [c for c in ("hito_1_val", "hito_2_val", "hito_3_val", "hito_5_val")
                 if c in df_descent.columns]
    aggs = []
    for hk in hito_cols:
        n = hk.split("_")[1]
        aggs.append(pl.col(hk).mean().round(1).alias(f"Hito {n} (días)"))
    if "Suspendidos" in df_descent.columns:
        aggs.append(pl.col("Suspendidos").sum().alias("Suspendidos"))
    if "Para cierre" in df_descent.columns:
        aggs.append(pl.col("Para cierre").sum().alias("Para cierre"))
    aggs.append(pl.len().alias("Total"))
    df_agr = df_descent.group_by("EJECUTOR").agg(aggs).sort("EJECUTOR").to_pandas()

    for r_i, row in enumerate(df_agr.to_dict(orient="records"), 5):
        bg = GRIS_ALT if r_i % 2 == 0 else BLANCO
        ws.row_dimensions[r_i].height = 22
        _data_cell(ws.cell(r_i, 1), row.get("EJECUTOR") or "", bg=bg, bold=True, color=AZUL_MED)
        col = 2
        for hk_label, hk_col in [("Hito 1 (días)", "hito_1_val"),
                                  ("Hito 2 (días)", "hito_2_val"),
                                  ("Hito 3 (días)", "hito_3_val"),
                                  ("Hito 5 (días)", "hito_5_val")]:
            v = row.get(hk_label)
            num = round(float(v), 1) if v is not None and str(v) != "nan" else None
            shown = round(num / 30.0, 1) if (hk_col == "hito_5_val" and num is not None) else num
            _data_cell(ws.cell(r_i, col), shown, bg=bg, center=True, fmt="#,##0.0")
            col += 1
            clasi = _clasificar_h(hk_col, num) if num is not None else None
            _sem_cell(ws.cell(r_i, col), clasi, bg_default=bg, hito_col=hk_col)
            col += 1
        for extra in ("Suspendidos", "Para cierre", "Total"):
            v = row.get(extra)
            iv = int(v) if v is not None and str(v) != "nan" else 0
            if extra == "Total":
                _data_cell(ws.cell(r_i, col), iv, bg="EFF6FF",
                           center=True, bold=True, color=AZUL_MED)
            else:
                _data_cell(ws.cell(r_i, col), iv, bg=bg, center=True)
            col += 1


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 5 · Detalle Descentralizadas
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_detalle_descent(wb, df_descent):
    ws = wb.create_sheet("Detalle Descentralizadas")
    ws.sheet_view.showGridLines = False

    base = [
        ("Ejecutor",            30, "EJECUTOR"),
        ("BPIN",                14, "BPIN"),
        ("Nombre del proyecto", 42, "NOMBRE DEL PROYECTO"),
        ("Estado proyecto",     20, "ESTADO PROYECTO"),
    ]
    if "ESTADO CONTRATO" in df_descent.columns:
        base.append(("Estado contrato", 16, "ESTADO CONTRATO"))
    if "AVANCE FÍSICO" in df_descent.columns:
        base.append(("Avance\nfísico",     11, "AVANCE FÍSICO"))
    if "AVANCE FINANCIERO" in df_descent.columns:
        base.append(("Avance\nfinanciero", 12, "AVANCE FINANCIERO"))
    for fcol, flbl, fw in [
        ("FECHA APROBACIÓN PROYECTO",          "Fecha\naprobación",        12),
        ("FECHA DE APERTURA DEL PRIMER PROCESO","Fecha apertura\nprimer proc.", 14),
        ("FECHA DE SUSCRIPCIÓN DEL CONTRATO PRINCIPAL",                  "Fecha\nsuscripción",       12),
        ("FECHA ACTA INICIO",                  "Fecha acta\ninicio",       12),
        ("HORIZONTE DEL PROYECTO",             "Horizonte\nproyecto",      12),
        ("FECHA DE CORTE GESPROY",             "Fecha corte\nGESPROY",     12),
    ]:
        if fcol in df_descent.columns:
            base.append((flbl, fw, fcol))
    # Descentralizadas tiene H1, H2, H3 y H5 — NO tiene H4 (requiere contratos)
    # ni H6 (sin fecha de finalización).
    for n, etiqueta in [(1, "H1"), (2, "H2"), (3, "H3"), (5, "H5")]:
        if f"hito_{n}_val" in df_descent.columns:
            base.append((f"{etiqueta}\ndías",       7, f"hito_{n}_val"))
            if f"clasi_{n}" in df_descent.columns:
                base.append((f"{etiqueta}\nalerta",     9, f"clasi_{n}"))
                base.append((f"{etiqueta} · Mensaje", 36, f"_msg_{n}"))

    cols = base
    NCOLS = len(cols)
    _title_row(ws,
               "Detalle de proyectos — Descentralizadas",
               f"Generado: {date.today().strftime('%d/%m/%Y')} · "
               "Fechas, días por hito (1-4) y nivel de alerta",
               NCOLS)
    ws.row_dimensions[3].height = 6
    for i, (label, w, _) in enumerate(cols, 1):
        _header_cell(ws.cell(4, i), label)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 50

    rows = df_descent.to_dicts()
    for r_i, row in enumerate(rows, 5):
        bg = GRIS_ALT if r_i % 2 == 0 else BLANCO
        ws.row_dimensions[r_i].height = 36
        msgs = {}
        for n in (1, 2, 3, 5):
            cv = row.get(f"clasi_{n}")
            cv_s = str(cv) if cv and str(cv) not in ("None", "nan", "") else None
            hk = HITO_BY_CLASI.get(f"clasi_{n}")
            if cv_s and hk in SEMAFOROS and cv_s in SEMAFOROS[hk]:
                msgs[f"_msg_{n}"] = SEMAFOROS[hk][cv_s][2]
            else:
                msgs[f"_msg_{n}"] = ""

        for c_i, (_, _, attr) in enumerate(cols, 1):
            cell = ws.cell(r_i, c_i)
            if attr in CLASI_COLS:
                _sem_cell(cell, row.get(attr), bg_default=bg, hito_col=HITO_BY_CLASI.get(attr))
            elif attr in HITO_COLS_DESCENT:
                v = row.get(attr)
                num = round(float(v), 1) if v is not None and str(v) != "nan" else None
                _data_cell(cell, num, bg=bg, center=True, fmt="#,##0.0")
            elif attr in PCT_COLS:
                pct = _avance_pct(row.get(attr))
                _data_cell(cell, round(pct, 1) if pct is not None else None,
                           bg=bg, center=True, fmt='0.0"%"')
            elif attr in DATE_COLS_DESCENT:
                _write_date(cell, row.get(attr), bg)
            elif attr.startswith("_msg"):
                _data_cell(cell, msgs.get(attr, ""), bg=bg, color="374151", italic=True)
            else:
                v = row.get(attr)
                _data_cell(cell, v if v else "—", bg=bg)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 7 · Detalle Municipios (sin hitos, sin contratos, sin evaluación)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_detalle_munic(wb, df_munic):
    ws = wb.create_sheet("Detalle Municipios")
    ws.sheet_view.showGridLines = False

    base = [
        ("Ejecutor (Municipio)", 32, "EJECUTOR"),
        ("BPIN",                 14, "BPIN"),
        ("Nombre del proyecto",  42, "NOMBRE DEL PROYECTO"),
    ]
    if "SECTOR" in df_munic.columns:
        base.append(("Sector", 22, "SECTOR"))
    base.append(("Estado proyecto", 20, "ESTADO PROYECTO"))
    if "ESTADO CONTRATO" in df_munic.columns:
        base.append(("Estado contrato", 16, "ESTADO CONTRATO"))
    if "AVANCE FÍSICO" in df_munic.columns:
        base.append(("Avance\nfísico",     11, "AVANCE FÍSICO"))
    if "AVANCE FINANCIERO" in df_munic.columns:
        base.append(("Avance\nfinanciero", 12, "AVANCE FINANCIERO"))

    cols = base
    NCOLS = len(cols)
    _title_row(ws,
               "Detalle de proyectos — Municipios",
               f"Generado: {date.today().strftime('%d/%m/%Y')} · "
               "Sin cálculo de hitos: la tabla no contiene fechas de seguimiento",
               NCOLS)
    ws.row_dimensions[3].height = 6
    for i, (label, w, _) in enumerate(cols, 1):
        _header_cell(ws.cell(4, i), label)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 32

    rows = df_munic.to_dicts()
    for r_i, row in enumerate(rows, 5):
        bg = GRIS_ALT if r_i % 2 == 0 else BLANCO
        ws.row_dimensions[r_i].height = 26
        for c_i, (_, _, attr) in enumerate(cols, 1):
            cell = ws.cell(r_i, c_i)
            if attr in PCT_COLS:
                pct = _avance_pct(row.get(attr))
                _data_cell(cell, round(pct, 1) if pct is not None else None,
                           bg=bg, center=True, fmt='0.0"%"')
            else:
                v = row.get(attr)
                _data_cell(cell, v if v else "—", bg=bg)


# ─────────────────────────────────────────────────────────────────────────────
# Hojas 8, 9 · Evaluación Sucre / Descentralizadas (mismo renderer)
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_eval(wb, df_eval, col_entidad, cols_calificacion,
                labels_calificacion, sheet_name, sub_titulo):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    headers = [col_entidad] + labels_calificacion + ["Promedio general"]
    widths  = [32] + [22] * len(labels_calificacion) + [20]
    NCOLS   = len(headers)
    _title_row(ws,
               f"Evaluación · {sub_titulo}",
               f"Generado: {date.today().strftime('%d/%m/%Y')} · "
               "Calificaciones promedio por entidad (escala 0–100)",
               NCOLS)
    ws.row_dimensions[3].height = 6
    for i, (lbl, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws.cell(4, i), lbl)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 38

    EVAL_SEM = [
        (80, VERDE_OSC, BLANCO),
        (60, "1754AB",  BLANCO),
        (40, "D88C16",  BLANCO),
        (0,  "E68878",  BLANCO),
    ]

    def _eval_fill(score):
        if score is None or str(score) in ("None", "nan", ""):
            return BLANCO, "1A2332"
        try:
            s = float(score)
        except Exception:
            return BLANCO, "1A2332"
        for thr, bg, fg in EVAL_SEM:
            if s >= thr:
                return bg, fg
        return BLANCO, "1A2332"

    if df_eval is None:
        ws.cell(5, 1, "Sin datos disponibles").font = _font(italic=True, color="9CA3AF")
        return

    df_pd = df_eval.to_pandas()
    if df_pd.empty:
        ws.cell(5, 1, "Sin datos disponibles").font = _font(italic=True, color="9CA3AF")
        return

    for r_i, row_vals in enumerate(df_pd.values.tolist(), 5):
        rd = dict(zip(df_pd.columns, row_vals))
        bg = GRIS_ALT if r_i % 2 == 0 else BLANCO
        ws.row_dimensions[r_i].height = 30
        ent = rd.get(col_entidad, "")
        _data_cell(ws.cell(r_i, 1), str(ent) if ent else "—",
                   bg=bg, bold=True, color=AZUL_MED)
        scores = []
        for c_i, col in enumerate(cols_calificacion, 2):
            v = rd.get(col)
            try:
                s = float(v) if v is not None and str(v) not in ("None", "nan", "") else None
            except Exception:
                s = None
            scores.append(s)
            cell = ws.cell(r_i, c_i)
            if s is not None:
                fill, color = _eval_fill(s)
                _data_cell(cell, round(s, 1), bg=fill, color=color,
                           center=True, bold=True, fmt="0.0")
            else:
                _data_cell(cell, "—", bg=bg, color="9CA3AF", center=True)
        valid = [s for s in scores if s is not None]
        prom = round(sum(valid) / len(valid), 1) if valid else None
        c_p = ws.cell(r_i, len(cols_calificacion) + 2)
        if prom is not None:
            fill, color = _eval_fill(prom)
            _data_cell(c_p, prom, bg=fill, color=color,
                       center=True, bold=True, fmt="0.0")
        else:
            _data_cell(c_p, "—", bg=bg, color="9CA3AF", center=True)


# ─────────────────────────────────────────────────────────────────────────────
# Función principal — orquesta la construcción del workbook
# ─────────────────────────────────────────────────────────────────────────────
def generar_excel(df_f=None, df_agr=None, clasi_por_entidad_map=None,
                  df_eval_sucre=None, cols_eval_sucre=None,
                  df_eval_desc=None, cols_eval_desc=None,
                  df_descent_hitos=None, df_municipios=None,
                  # alias retro-compatible
                  df_f_full=None):
    """
    Genera un reporte Excel global y consolidado, independiente de la vista
    activa en la app. Devuelve los bytes del archivo .xlsx.

    Hojas potenciales (cada una se omite si los datos no están disponibles):
      1. Resumen Departamento
      2. Detalle Departamento
      3. Reporte Semanal Dpto
      4. Resumen Descentralizadas
      5. Detalle Descentralizadas
      6. Reporte Semanal Descent
      7. Detalle Municipios
      8. Evaluación Sucre
      9. Evaluación Descentralizadas
    """
    if df_f is None and df_f_full is not None:
        df_f = df_f_full

    wb = Workbook()

    # ── DEPARTAMENTO ──────────────────────────────────────────────────────
    if df_f is not None and df_agr is not None:
        _sheet_resumen_dpto(wb, df_agr)
        _sheet_detalle_dpto(wb, df_f)
        _sheet_reporte_semanal(wb, df_f, "ENTIDAD O SECRETARIA",
                               "Dependencia / Secretaría",
                               "Reporte Semanal Dpto",
                               "Departamento de Sucre")
    else:
        # Si no hay datos de Departamento, dejamos al menos una hoja vacía
        # como placeholder (el primer create_sheet promueve esa hoja a activa).
        ws = wb.active
        ws.title = "Sin datos"
        ws.cell(1, 1, "No se cargaron datos del Departamento.").font = (
            _font(italic=True, color="9CA3AF")
        )

    # ── DESCENTRALIZADAS ──────────────────────────────────────────────────
    if df_descent_hitos is not None and df_descent_hitos.height > 0:
        _sheet_resumen_descent(wb, df_descent_hitos)
        _sheet_detalle_descent(wb, df_descent_hitos)
        _sheet_reporte_semanal(wb, df_descent_hitos, "EJECUTOR",
                               "Ejecutor",
                               "Reporte Semanal Descent",
                               "Descentralizadas")

    # ── MUNICIPIOS ────────────────────────────────────────────────────────
    if df_municipios is not None and df_municipios.height > 0:
        _sheet_detalle_munic(wb, df_municipios)

    # ── EVALUACIONES ──────────────────────────────────────────────────────
    eval_labels_map = dict(zip(COLS_EVAL, COLS_EVAL_LABELS))
    if df_eval_sucre is not None and cols_eval_sucre:
        labels_s = [eval_labels_map.get(c, c) for c in cols_eval_sucre]
        _sheet_eval(wb, df_eval_sucre, "ENTIDAD O SECRETARIA",
                    cols_eval_sucre, labels_s,
                    "Evaluación Sucre", "Departamento de Sucre")
    if df_eval_desc is not None and cols_eval_desc:
        labels_d = [eval_labels_map.get(c, c) for c in cols_eval_desc]
        _sheet_eval(wb, df_eval_desc, "EJECUTOR",
                    cols_eval_desc, labels_d,
                    "Evaluación Descentralizadas", "Entidades Descentralizadas")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
