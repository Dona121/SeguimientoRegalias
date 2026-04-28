"""
export.py
Generación del reporte Excel con todas las hojas:
Resumen por Entidad, Detalle Proyectos, Evaluación Sucre, Evaluación Descentralizadas.
"""
from constants import (
    INTERVALOS, COLS_EVAL, COLS_EVAL_LABELS, SEMAFOROS, C,
)
import streamlit as st
import polars as pl
import pandas as pd
import io
import html
import json
import logging
import urllib.parse
import urllib.request
import streamlit.components.v1 as components
import datetime as _dt
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

_log = logging.getLogger(__name__)


def generar_excel(df_f=None, df_agr=None, clasi_por_entidad_map=None,
                  df_eval_sucre=None, cols_eval_sucre=None,
                  df_eval_desc=None, cols_eval_desc=None,
                  df_descent_hitos=None, df_municipios=None,
                  # Aliases retro-compatibles
                  df_f_full=None):
    """
    Genera reporte Excel global y consolidado, independiente de la vista activa
    en la app. Hojas potenciales:
      1. Resumen Departamento por Entidad
      2. Detalle Proyectos Departamento (con fechas y alertas)
      3. Resumen Descentralizadas por Ejecutor (hitos 1-4)
      4. Detalle Proyectos Descentralizadas
      5. Detalle Proyectos Municipios (sin hitos)
      6. Evaluación Sucre
      7. Evaluación Descentralizadas

    Parámetros:
      df_f  / df_f_full     : pl.DataFrame del Departamento con fechas + hitos.
      df_agr                : agrupado por ENTIDAD O SECRETARIA con promedios.
      clasi_por_entidad_map : dict {entidad -> {clasi_k -> valor_mas_frecuente}}.
      df_descent_hitos      : pl.DataFrame de Descentralizadas con hitos 1-4.
      df_municipios         : pl.DataFrame de Municipios (proyectos básicos).
    """
    # Compat: el nombre antiguo era df_f_full
    if df_f is None and df_f_full is not None:
        df_f = df_f_full
    df_f_full = df_f  # alias usado más abajo en el código existente
    # ── Paleta ────────────────────────────────────────────────────────────────
    AZUL_OSC  = "003D6C"
    VERDE_OSC = "005931"
    AZUL_MED  = "1754AB"
    BLANCO    = "FFFFFF"
    GRIS_ALT  = "F7FAFD"
    GRIS_BRD  = "D1D5DB"

    # ── Semáforo Excel — actualizado con claves de hito 3 nuevas ─────────────
    SEM_FILL = {
        # Hito 1 y 2 y 5
        "0-100":   ("D1FAE5", "065F46"),
        "101-150": ("FEF3C7", "92400E"),
        "151-180": ("FFEDD5", "9A3412"),
        ">180":    ("1E293B", "F1F5F9"),
        # Hito 3 — actualizado
        "0-15":    ("D1FAE5", "065F46"),
        "16-30":   ("FEF3C7", "92400E"),
        "31-45":   ("FFEDD5", "9A3412"),
        ">45":     ("1E293B", "F1F5F9"),
        # Hito 4
        "0-1":     ("D1FAE5", "065F46"),
        "1.1-3":   ("FEF3C7", "92400E"),
        "3.1-6":   ("FFEDD5", "9A3412"),
        ">6":      ("1E293B", "F1F5F9"),
    }
    SEM_NOMBRE = {
        # Hito 1, 2, 5
        "0-100":   "Verde",
        "101-150": "Naranja",
        "151-180": "Rojo",
        ">180":    "Negro",
        # Hito 3 — actualizado
        "0-15":    "Verde",
        "16-30":   "Naranja",
        "31-45":   "Rojo",
        ">45":     "Negro",
        # Hito 4
        "0-1":     "Verde",
        "1.1-3":   "Naranja",
        "3.1-6":   "Rojo",
        ">6":      "Negro",
    }

    # Mensajes completos desde SEMAFOROS global
    SEM_MSG = {}
    for hk, vals in SEMAFOROS.items():
        for label, (_, color_nombre, msg) in vals.items():
            SEM_MSG[label] = msg

    def _side():   return Side(style="thin", color=GRIS_BRD)
    def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    def _font(bold=False, color="1A2332", size=11, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)
    def _fill(color): return PatternFill("solid", fgColor=color)
    def _align(h="left", wrap=True, v="top"):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _header_cell(cell, text):
        cell.value      = text
        cell.font       = Font(name="Calibri", bold=True, color=BLANCO, size=11)
        cell.fill       = _fill(AZUL_OSC)
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = _border()

    def _data_cell(cell, value, bg=BLANCO, bold=False, color="1A2332", center=False, fmt=None):
        cell.value      = value
        cell.font       = _font(bold=bold, color=color)
        cell.fill       = _fill(bg)
        cell.alignment  = _align("center" if center else "left")
        cell.border     = _border()
        if fmt: cell.number_format = fmt

    def _sem_cell(cell, clasi, bg_row=BLANCO):
        clasi_s = str(clasi) if clasi and str(clasi) not in ("nan", "None", "") else None
        if clasi_s and clasi_s in SEM_FILL:
            bg, fg = SEM_FILL[clasi_s]
            cell.value     = clasi_s
            cell.font      = Font(name="Calibri", bold=True, color=fg, size=11)
            cell.fill      = _fill(bg)
            cell.alignment = _align("center")
            cell.border    = _border()
            msg    = SEM_MSG.get(clasi_s, "")
            nombre = SEM_NOMBRE.get(clasi_s, "")
            if msg:
                cell.comment = Comment(f"● {nombre}\n{msg}", "Sistema", height=60, width=300)
        else:
            _data_cell(cell, "—", bg=bg_row, center=True, color="9CA3AF")

    def _title_row(ws, text, sub, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(1, 1, text)
        c.font      = Font(name="Calibri", bold=True, size=14, color=BLANCO)
        c.fill      = _fill(AZUL_OSC)
        c.alignment = _align("left")
        ws.row_dimensions[1].height = 34

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c2 = ws.cell(2, 1, sub)
        c2.font      = Font(name="Calibri", size=11, color=BLANCO, italic=True)
        c2.fill      = _fill(VERDE_OSC)
        c2.alignment = _align("left")
        ws.row_dimensions[2].height = 20

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1 · Resumen por Entidad
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen por Entidad"
    ws1.sheet_view.showGridLines = False

    H1_COLS = [
        ("Entidad / Secretaría",                                   28),
        ("H1 · Sin contratar\nSin apertura\nPromedio días",        13),
        ("H1\nAlerta",                                              9),
        ("H2 · Sin contratar\nCon apertura\nPromedio días",        13),
        ("H2\nAlerta",                                              9),
        ("H3 · Contratado\nSin acta inicio\nPromedio días",        13),
        ("H3\nAlerta",                                              9),
        ("H4 · En ejecución\nRezagado\nPromedio meses",            13),
        ("H4\nAlerta",                                              9),
        ("H5 · Terminados\nPromedio días",                         13),
        ("H5\nAlerta",                                              9),
        ("Suspendidos",                                            12),
        ("Para cierre",                                            12),
        ("Total\nproyectos",                                       10),
    ]
    NCOLS1 = len(H1_COLS)
    _title_row(
        ws1,
        "Seguimiento y Evaluación · Regalías — Resumen por Entidad",
        f"Generado: {date.today().strftime('%d/%m/%Y')}   ·   Promedio de días/meses por hito   ·   Los colores indican el nivel de alerta",
        NCOLS1,
    )

    ws1.row_dimensions[3].height = 6  # separador visual

    for ci, (label, width) in enumerate(H1_COLS, 1):
        _header_cell(ws1.cell(4, ci), label)
        ws1.column_dimensions[get_column_letter(ci)].width = width
    ws1.row_dimensions[4].height = 48

    agr_pd   = df_agr.to_pandas()
    AGR_COLS = list(agr_pd.columns)

    HITO_AGR = [
        ("Hito 1 (días)", "clasi_1"),
        ("Hito 2 (días)", "clasi_2"),
        ("Hito 3 (días)", "clasi_3"),
        ("Hito 4 (días)", "clasi_4"),
        ("Hito 5 (días)", "clasi_5"),
    ]

    for ri, row_vals in enumerate(agr_pd.values.tolist(), 5):
        row_dict = dict(zip(AGR_COLS, row_vals))
        bg      = GRIS_ALT if ri % 2 == 0 else BLANCO
        entidad = row_dict.get("ENTIDAD O SECRETARIA") or ""
        ws1.row_dimensions[ri].height = 24

        _data_cell(ws1.cell(ri, 1), entidad, bg=bg, bold=True, color=AZUL_MED)

        col = 2
        for dias_col, clasi_key in HITO_AGR:
            dias     = row_dict.get(dias_col)
            dias_num = round(float(dias), 1) if dias is not None and str(dias) != "nan" else None
            # H4 — el encabezado de la columna está en meses, así que el valor
            # mostrado también va en meses (los demás hitos van en días).
            if clasi_key == "clasi_4" and dias_num is not None:
                valor_mostrado = round(dias_num / 30.0, 1)
            else:
                valor_mostrado = dias_num
            _data_cell(ws1.cell(ri, col), valor_mostrado, bg=bg, center=True, fmt="#,##0.0")
            col += 1

            # Clasificar desde el promedio — coherente con el valor mostrado
            if dias_num is not None:
                hito_col = {
                    "clasi_1": "hito_1_val", "clasi_2": "hito_2_val",
                    "clasi_3": "hito_3_val", "clasi_4": "hito_4_val",
                    "clasi_5": "hito_5_val",
                }.get(clasi_key)
                clasi_excel = None
                if hito_col == "hito_4_val":
                    m = dias_num / 30.0
                    clasi_excel = "0-1" if m <= 1 else "1.1-3" if m <= 3 else "3.1-6" if m <= 6 else ">6"
                else:
                    for label, lo, hi in INTERVALOS.get(hito_col, []):
                        if (hi is None and dias_num >= lo) or (hi is not None and lo <= dias_num <= hi):
                            clasi_excel = label
                            break
            else:
                clasi_excel = None
            _sem_cell(ws1.cell(ri, col), clasi_excel, bg_row=bg)
            col += 1

        for extra_col, bold_it, extra_bg in [
            ("Suspendidos", False, None),
            ("Para cierre", False, None),
            ("Total",       True,  "EFF6FF"),
        ]:
            val = row_dict.get(extra_col)
            v   = int(val) if val is not None and str(val) != "nan" else 0
            c   = ws1.cell(ri, col)
            _data_cell(c, v,
                       bg=extra_bg or bg,
                       bold=bold_it,
                       color=AZUL_MED if extra_bg else "1A2332",
                       center=True)
            col += 1

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 2 · Detalle por Proyecto
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detalle Proyectos")
    ws2.sheet_view.showGridLines = False

    H2_COLS = [
        ("Entidad /\nSecretaría",              26, "ENTIDAD O SECRETARIA"),
        ("BPIN",                               13, "BPIN"),
        ("Nombre del proyecto",                42, "NOMBRE PROYECTO"),
        ("Estado\nproyecto",                   18, "ESTADO PROYECTO"),
        ("Estado\ncontrato",                   18, "ESTADO CONTRATO"),
        ("Fecha\naprobación",                  13, "FECHA APROBACIÓN PROYECTO"),
        ("Fecha apertura\nprimer proceso",     13, "FECHA DE APERTURA DEL PRIMER PROCESO"),
        ("Fecha\nsuscripción",                 13, "FECHA SUSCRIPCION"),
        ("Fecha acta\ninicio",                 13, "FECHA ACTA INICIO"),
        ("Horizonte\nproyecto",                13, "HORIZONTE DEL PROYECTO"),
        ("Fecha\nfinalización",                13, "FECHA DE FINALIZACIÓN"),
        ("Fecha corte\nGESPROY",               13, "FECHA DE CORTE GESPROY"),
        ("H1\ndías",                            8, "hito_1_val"),
        ("H1\nalerta",                          9, "clasi_1"),
        ("H1 · Mensaje de alerta",             34, "_msg_1"),
        ("H2\ndías",                            8, "hito_2_val"),
        ("H2\nalerta",                          9, "clasi_2"),
        ("H2 · Mensaje de alerta",             34, "_msg_2"),
        ("H3\ndías",                            8, "hito_3_val"),
        ("H3\nalerta",                          9, "clasi_3"),
        ("H3 · Mensaje de alerta",             34, "_msg_3"),
        ("H4\ndías",                            8, "hito_4_val"),
        ("H4\nalerta",                          9, "clasi_4"),
        ("H4 · Mensaje de alerta",             34, "_msg_4"),
        ("H5\ndías",                            8, "hito_5_val"),
        ("H5\nalerta",                          9, "clasi_5"),
        ("H5 · Mensaje de alerta",             34, "_msg_5"),
        ("Suspendido",                         11, "Suspendidos"),
        ("Para\ncierre",                       10, "Para cierre"),
    ]
    NCOLS2 = len(H2_COLS)
    _title_row(
        ws2,
        "Seguimiento y Evaluación · Regalías — Detalle por Proyecto",
        f"Generado: {date.today().strftime('%d/%m/%Y')}   ·   Incluye fechas usadas en el cálculo de hitos y niveles de alerta con mensajes",
        NCOLS2,
    )

    ws2.row_dimensions[3].height = 6

    for ci, (label, width, _) in enumerate(H2_COLS, 1):
        _header_cell(ws2.cell(3, ci), label)
        ws2.column_dimensions[get_column_letter(ci)].width = width
    ws2.row_dimensions[3].height = 48

    DATE_COLS = {
        "FECHA APROBACIÓN PROYECTO", "FECHA DE APERTURA DEL PRIMER PROCESO",
        "FECHA SUSCRIPCION", "FECHA ACTA INICIO", "HORIZONTE DEL PROYECTO",
        "FECHA DE FINALIZACIÓN", "FECHA DE CORTE GESPROY",
    }
    HITO_CLASI_PAIRS = [
        ("hito_1_val", "clasi_1"), ("hito_2_val", "clasi_2"),
        ("hito_3_val", "clasi_3"), ("hito_4_val", "clasi_4"),
        ("hito_5_val", "clasi_5"),
    ]
    NUM_COLS  = {"hito_1_val", "hito_2_val", "hito_3_val", "hito_4_val", "hito_5_val"}
    FLAG_COLS = {"Suspendidos", "Para cierre"}

    rows = df_f_full.to_dicts()
    for ri2, row in enumerate(rows, 4):
        bg = GRIS_ALT if ri2 % 2 == 0 else BLANCO
        ws2.row_dimensions[ri2].height = 80

        msgs = {}
        for n, (_, clasi_col) in enumerate(HITO_CLASI_PAIRS, 1):
            cv   = row.get(clasi_col)
            cv_s = str(cv) if cv and str(cv) not in ("nan", "None", "") else None
            msgs[f"_msg_{n}"] = SEM_MSG.get(cv_s, "") if cv_s else ""

        for ci, (_, _, attr) in enumerate(H2_COLS, 1):
            cell = ws2.cell(ri2, ci)
            val  = row.get(attr) if not attr.startswith("_msg") else msgs.get(attr, "")

            if attr in ("clasi_1", "clasi_2", "clasi_3", "clasi_4", "clasi_5"):
                _sem_cell(cell, val, bg_row=bg)

            elif attr in NUM_COLS:
                v = round(float(val), 1) if val is not None and str(val) != "nan" else None
                _data_cell(cell, v, bg=bg, center=True, color="1A2332", fmt="#,##0.0")

            elif attr in DATE_COLS:
                if val is not None and str(val) not in ("nan", "NaT", "None", ""):
                    if isinstance(val, (_dt.date, _dt.datetime)):
                        cell.value         = val
                        cell.number_format = "DD/MM/YYYY"
                        cell.font          = _font()
                        cell.fill          = _fill(bg)
                        cell.alignment     = _align("center")
                        cell.border        = _border()
                    else:
                        _data_cell(cell, str(val), bg=bg, center=True)
                else:
                    _data_cell(cell, "—", bg=bg, center=True, color="9CA3AF")

            elif attr in FLAG_COLS:
                es_si = val is not None and str(val) not in ("nan", "None", "0", "") and val != 0
                _data_cell(
                    cell, "Sí" if es_si else "No",
                    bg="FEF3C7" if es_si else bg,
                    center=True, bold=es_si,
                    color="92400E" if es_si else "6B7280",
                )

            elif attr.startswith("_msg"):
                _data_cell(cell, val, bg=bg, color="374151")
                cell.font = Font(name="Calibri", size=11, color="374151", italic=True)

            else:
                _data_cell(cell, val if val else "—", bg=bg)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 3 · Evaluación del modelo
    # ══════════════════════════════════════════════════════════════════════════
    def _ws_eval(ws, df_eval, col_entidad, cols_calificacion, labels_calificacion, subtitulo):
        ws.sheet_view.showGridLines = False

        headers = [col_entidad] + labels_calificacion + ["Promedio general"]
        widths  = [30] + [22] * len(labels_calificacion) + [18]
        NCOLS   = len(headers)

        _title_row(
            ws,
            f"Seguimiento y Evaluación · Regalías — {subtitulo}",
            f"Generado: {date.today().strftime('%d/%m/%Y')}   ·   Calificaciones promedio por entidad (escala 0–100)",
            NCOLS,
        )

        ws.row_dimensions[3].height = 6

        for ci, (label, width) in enumerate(zip(headers, widths), 1):
            _header_cell(ws.cell(3, ci), label)
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[3].height = 48

        EVAL_SEM = [
            (80, VERDE_OSC, "FFFFFF"),
            (60, "1754AB",  "FFFFFF"),
            (40, "D88C16",  "FFFFFF"),
            (0,  "E68878",  "FFFFFF"),
        ]

        def _eval_fill(score):
            if score is None or str(score) in ("nan", "None", ""):
                return BLANCO, "1A2332"
            s = float(score)
            for threshold, bg, fg in EVAL_SEM:
                if s >= threshold:
                    return bg, fg
            return BLANCO, "1A2332"

        df_pd = df_eval.to_pandas() if df_eval is not None else None
        if df_pd is None or df_pd.empty:
            ws.cell(4, 1, "Sin datos disponibles").font = Font(
                name="Calibri", size=11, italic=True, color="9CA3AF"
            )
            return

        for ri, row_vals in enumerate(df_pd.values.tolist(), 4):
            row_dict = dict(zip(df_pd.columns, row_vals))
            bg = GRIS_ALT if ri % 2 == 0 else BLANCO
            ws.row_dimensions[ri].height = 30

            entidad_val = row_dict.get(col_entidad, "")
            _data_cell(ws.cell(ri, 1),
                       str(entidad_val) if entidad_val else "—",
                       bg=bg, bold=True, color=AZUL_MED)

            scores = []
            for ci, col in enumerate(cols_calificacion, 2):
                val   = row_dict.get(col)
                score = float(val) if val is not None and str(val) not in ("nan", "None", "") else None
                scores.append(score)
                cell  = ws.cell(ri, ci)
                if score is not None:
                    bg_s, fg_s = _eval_fill(score)
                    _data_cell(cell, round(score, 1), bg=bg_s, color=fg_s, center=True, fmt="0.0")
                else:
                    _data_cell(cell, "—", bg=bg, color="9CA3AF", center=True)

            valid  = [s for s in scores if s is not None]
            prom   = round(sum(valid) / len(valid), 1) if valid else None
            ci_prom = len(cols_calificacion) + 2
            cell_p  = ws.cell(ri, ci_prom)
            if prom is not None:
                bg_p, fg_p = _eval_fill(prom)
                _data_cell(cell_p, prom, bg=bg_p, color=fg_p, center=True, bold=True, fmt="0.0")
            else:
                _data_cell(cell_p, "—", bg=bg, color="9CA3AF", center=True)

    COLS_EVAL_LABELS_MAP = dict(zip(COLS_EVAL, COLS_EVAL_LABELS))

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA · Resumen Descentralizadas (hitos 1-4 promediados por ejecutor)
    # ══════════════════════════════════════════════════════════════════════════
    if df_descent_hitos is not None and df_descent_hitos.height > 0:
        ws_d_res = wb.create_sheet("Resumen Descentralizadas")
        ws_d_res.sheet_view.showGridLines = False

        H_DRES_COLS = [
            ("Ejecutor",                                       30),
            ("H1 · Sin contratar\nSin apertura\nPromedio días", 13),
            ("H1\nAlerta",                                       9),
            ("H2 · Sin contratar\nCon apertura\nPromedio días", 13),
            ("H2\nAlerta",                                       9),
            ("H3 · Contratado\nSin acta inicio\nPromedio días", 13),
            ("H3\nAlerta",                                       9),
            ("H4 · En ejecución\nRezagado\nPromedio meses",     13),
            ("H4\nAlerta",                                       9),
            ("Suspendidos",                                     12),
            ("Para cierre",                                     12),
            ("Total\nproyectos",                                10),
        ]
        NCOLS_D = len(H_DRES_COLS)
        _title_row(ws_d_res,
                   "Seguimiento · Regalías — Resumen Descentralizadas",
                   f"Generado: {date.today().strftime('%d/%m/%Y')}   ·   "
                   "Promedio de días/meses por hito (H5 no aplica para esta tabla)",
                   NCOLS_D)
        ws_d_res.row_dimensions[3].height = 6
        for ci, (lbl, w) in enumerate(H_DRES_COLS, 1):
            _header_cell(ws_d_res.cell(4, ci), lbl)
            ws_d_res.column_dimensions[get_column_letter(ci)].width = w
        ws_d_res.row_dimensions[4].height = 48

        # Agregar promedios por EJECUTOR
        _hito_present_d = [c for c in ("hito_1_val","hito_2_val","hito_3_val","hito_4_val")
                           if c in df_descent_hitos.columns]
        _agg_d = []
        for hk in _hito_present_d:
            n = hk.split("_")[1]
            _agg_d.append(pl.col(hk).mean().round(1).alias(f"Hito {n} (días)"))
        if "Suspendidos" in df_descent_hitos.columns:
            _agg_d.append(pl.col("Suspendidos").sum().alias("Suspendidos"))
        if "Para cierre" in df_descent_hitos.columns:
            _agg_d.append(pl.col("Para cierre").sum().alias("Para cierre"))
        _agg_d.append(pl.len().alias("Total"))
        agr_d = (df_descent_hitos.group_by("EJECUTOR").agg(_agg_d)
                 .sort("EJECUTOR").to_pandas())

        for ri, row_vals in enumerate(agr_d.values.tolist(), 5):
            row_d = dict(zip(agr_d.columns, row_vals))
            bg    = GRIS_ALT if ri % 2 == 0 else BLANCO
            ws_d_res.row_dimensions[ri].height = 24
            _data_cell(ws_d_res.cell(ri, 1), row_d.get("EJECUTOR") or "",
                       bg=bg, bold=True, color=AZUL_MED)
            col = 2
            for hk_label, clasi_key in [("Hito 1 (días)","clasi_1"), ("Hito 2 (días)","clasi_2"),
                                        ("Hito 3 (días)","clasi_3"), ("Hito 4 (días)","clasi_4")]:
                v = row_d.get(hk_label)
                v_num = round(float(v), 1) if v is not None and str(v) != "nan" else None
                if clasi_key == "clasi_4" and v_num is not None:
                    v_show = round(v_num / 30.0, 1)
                else:
                    v_show = v_num
                _data_cell(ws_d_res.cell(ri, col), v_show, bg=bg, center=True, fmt="#,##0.0")
                col += 1
                # Clasificar
                clasi_excel = None
                if v_num is not None:
                    if clasi_key == "clasi_4":
                        m = v_num / 30.0
                        clasi_excel = "0-1" if m <= 1 else "1.1-3" if m <= 3 else "3.1-6" if m <= 6 else ">6"
                    else:
                        hk_col = {"clasi_1":"hito_1_val","clasi_2":"hito_2_val","clasi_3":"hito_3_val"}[clasi_key]
                        for label, lo, hi in INTERVALOS.get(hk_col, []):
                            if (hi is None and v_num >= lo) or (hi is not None and lo <= v_num <= hi):
                                clasi_excel = label
                                break
                _sem_cell(ws_d_res.cell(ri, col), clasi_excel, bg_row=bg)
                col += 1
            for extra_col in ("Suspendidos","Para cierre"):
                v = row_d.get(extra_col)
                _data_cell(ws_d_res.cell(ri, col),
                           int(v) if v is not None and str(v) != "nan" else 0,
                           bg=bg, center=True)
                col += 1
            _data_cell(ws_d_res.cell(ri, col),
                       int(row_d.get("Total") or 0),
                       bg="EFF6FF", center=True, bold=True, color=AZUL_MED)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA · Detalle Descentralizadas (proyectos con sus hitos y alertas)
    # ══════════════════════════════════════════════════════════════════════════
    if df_descent_hitos is not None and df_descent_hitos.height > 0:
        ws_d_det = wb.create_sheet("Detalle Descentralizadas")
        ws_d_det.sheet_view.showGridLines = False

        # Construir columnas dinámicamente según lo que esté presente
        H_DET_BASE = [
            ("Ejecutor",                  28, "EJECUTOR"),
            ("BPIN",                      13, "BPIN"),
            ("Nombre del proyecto",       42, "NOMBRE DEL PROYECTO"),
            ("Estado proyecto",           18, "ESTADO PROYECTO"),
        ]
        if "ESTADO CONTRATO" in df_descent_hitos.columns:
            H_DET_BASE.append(("Estado contrato", 18, "ESTADO CONTRATO"))
        if "AVANCE FÍSICO" in df_descent_hitos.columns:
            H_DET_BASE.append(("Avance físico", 11, "AVANCE FÍSICO"))
        if "AVANCE FINANCIERO" in df_descent_hitos.columns:
            H_DET_BASE.append(("Avance financiero", 13, "AVANCE FINANCIERO"))
        # Fechas que sí están
        for fcol, flbl, fw in [
            ("FECHA APROBACIÓN PROYECTO",         "Fecha\naprobación",       13),
            ("FECHA DE APERTURA DEL PRIMER PROCESO","Fecha apertura\nprimer proc.", 14),
            ("FECHA SUSCRIPCION",                 "Fecha\nsuscripción",      13),
            ("FECHA ACTA INICIO",                 "Fecha acta\ninicio",      13),
            ("HORIZONTE DEL PROYECTO",            "Horizonte\nproyecto",     13),
            ("FECHA DE CORTE GESPROY",            "Fecha corte\nGESPROY",    13),
        ]:
            if fcol in df_descent_hitos.columns:
                H_DET_BASE.append((flbl, fw, fcol))
        # Hitos
        for hk_label, hk_col in [("H1 días","hito_1_val"), ("H2 días","hito_2_val"),
                                 ("H3 días","hito_3_val"), ("H4 días","hito_4_val")]:
            if hk_col in df_descent_hitos.columns:
                H_DET_BASE.append((hk_label, 8, hk_col))
                clasi_col = "clasi_" + hk_col.split("_")[1]
                if clasi_col in df_descent_hitos.columns:
                    H_DET_BASE.append((hk_label.replace("días","alerta"), 9, clasi_col))

        NCOLS_DD = len(H_DET_BASE)
        _title_row(ws_d_det,
                   "Seguimiento · Regalías — Detalle Descentralizadas",
                   f"Generado: {date.today().strftime('%d/%m/%Y')}   ·   Proyectos con sus hitos y alertas",
                   NCOLS_DD)
        ws_d_det.row_dimensions[3].height = 6
        for ci, (lbl, w, _) in enumerate(H_DET_BASE, 1):
            _header_cell(ws_d_det.cell(4, ci), lbl)
            ws_d_det.column_dimensions[get_column_letter(ci)].width = w
        ws_d_det.row_dimensions[4].height = 48

        DATE_SET_D = {
            "FECHA APROBACIÓN PROYECTO","FECHA DE APERTURA DEL PRIMER PROCESO",
            "FECHA SUSCRIPCION","FECHA ACTA INICIO","HORIZONTE DEL PROYECTO",
            "FECHA DE CORTE GESPROY",
        }
        NUM_SET_D = {"hito_1_val","hito_2_val","hito_3_val","hito_4_val",
                     "AVANCE FÍSICO","AVANCE FINANCIERO"}

        rows_d = df_descent_hitos.to_dicts()
        for ri2, row in enumerate(rows_d, 5):
            bg = GRIS_ALT if ri2 % 2 == 0 else BLANCO
            ws_d_det.row_dimensions[ri2].height = 32
            for ci, (_, _, attr) in enumerate(H_DET_BASE, 1):
                cell = ws_d_det.cell(ri2, ci)
                val  = row.get(attr)
                if attr.startswith("clasi_"):
                    _sem_cell(cell, val, bg_row=bg)
                elif attr in NUM_SET_D:
                    v = round(float(val), 1) if val is not None and str(val) != "nan" else None
                    _data_cell(cell, v, bg=bg, center=True, fmt="#,##0.0")
                elif attr in DATE_SET_D:
                    if val is not None and str(val) not in ("nan","NaT","None",""):
                        if isinstance(val, (_dt.date, _dt.datetime)):
                            cell.value = val
                            cell.number_format = "DD/MM/YYYY"
                            cell.font = _font()
                            cell.fill = _fill(bg)
                            cell.alignment = _align("center")
                            cell.border = _border()
                        else:
                            _data_cell(cell, str(val), bg=bg, center=True)
                    else:
                        _data_cell(cell, "—", bg=bg, center=True, color="9CA3AF")
                else:
                    _data_cell(cell, val if val else "—", bg=bg)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA · Detalle Municipios (sin hitos — la tabla no tiene fechas suficientes)
    # ══════════════════════════════════════════════════════════════════════════
    if df_municipios is not None and df_municipios.height > 0:
        ws_m = wb.create_sheet("Detalle Municipios")
        ws_m.sheet_view.showGridLines = False

        H_M_COLS = [
            ("Ejecutor (Municipio)", 28, "EJECUTOR"),
            ("BPIN",                 13, "BPIN"),
            ("Nombre del proyecto",  42, "NOMBRE DEL PROYECTO"),
            ("Estado proyecto",      18, "ESTADO PROYECTO"),
        ]
        if "ESTADO CONTRATO" in df_municipios.columns:
            H_M_COLS.append(("Estado contrato", 18, "ESTADO CONTRATO"))
        if "AVANCE FÍSICO" in df_municipios.columns:
            H_M_COLS.append(("Avance físico", 11, "AVANCE FÍSICO"))
        if "AVANCE FINANCIERO" in df_municipios.columns:
            H_M_COLS.append(("Avance financiero", 13, "AVANCE FINANCIERO"))

        NCOLS_M = len(H_M_COLS)
        _title_row(ws_m,
                   "Seguimiento · Regalías — Detalle Municipios",
                   f"Generado: {date.today().strftime('%d/%m/%Y')}   ·   "
                   "Listado de proyectos por municipio (sin cálculo de hitos)",
                   NCOLS_M)
        ws_m.row_dimensions[3].height = 6
        for ci, (lbl, w, _) in enumerate(H_M_COLS, 1):
            _header_cell(ws_m.cell(4, ci), lbl)
            ws_m.column_dimensions[get_column_letter(ci)].width = w
        ws_m.row_dimensions[4].height = 36

        rows_m = df_municipios.to_dicts()
        NUM_SET_M = {"AVANCE FÍSICO","AVANCE FINANCIERO"}
        for ri2, row in enumerate(rows_m, 5):
            bg = GRIS_ALT if ri2 % 2 == 0 else BLANCO
            ws_m.row_dimensions[ri2].height = 24
            for ci, (_, _, attr) in enumerate(H_M_COLS, 1):
                cell = ws_m.cell(ri2, ci)
                val = row.get(attr)
                if attr in NUM_SET_M:
                    v = round(float(val), 1) if val is not None and str(val) != "nan" else None
                    _data_cell(cell, v, bg=bg, center=True, fmt="#,##0.0")
                else:
                    _data_cell(cell, val if val else "—", bg=bg)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJAS · Evaluaciones (Sucre y Descentralizadas)
    # ══════════════════════════════════════════════════════════════════════════
    if df_eval_sucre is not None and cols_eval_sucre:
        ws3    = wb.create_sheet("Evaluación Sucre")
        labels_s = [COLS_EVAL_LABELS_MAP.get(c, c) for c in cols_eval_sucre]
        _ws_eval(ws3, df_eval_sucre, "ENTIDAD O SECRETARIA",
                 cols_eval_sucre, labels_s, "Evaluación · Departamento de Sucre")

    if df_eval_desc is not None and cols_eval_desc:
        ws4    = wb.create_sheet("Evaluación Descentralizadas")
        labels_d = [COLS_EVAL_LABELS_MAP.get(c, c) for c in cols_eval_desc]
        _ws_eval(ws4, df_eval_desc, "EJECUTOR",
                 cols_eval_desc, labels_d, "Evaluación · Entidades Descentralizadas")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
